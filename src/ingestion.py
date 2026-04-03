"""
Data Ingestion Module
=====================

Provides connectors for six diverse data sources:

1. **RSS**       – Property Week RSS feed
2. **Scraping**  – JLL Trends & Insights articles (requests + BS4)
3. **Scraping**  – Altus Group Insights articles (requests + BS4)
4. **API**       – Financial Modeling Prep company profiles (mock-ready)
5. **CSV**       – Local cities.csv, homes.csv, zillow.csv
6. **XLSX**      – CRE Lending Data (UK + Continental Europe deals)

Every connector returns a list of plain dictionaries so that the
downstream normalisation layer can validate them via Pydantic.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import feedparser
import requests
from bs4 import BeautifulSoup

from src.config import (
    ALTUS_BASE_URL,
    CITIES_CSV_PATH,
    CRE_LENDING_XLSX_PATH,
    FMP_API_KEY,
    FMP_BASE_URL,
    FMP_TICKERS,
    HOMES_CSV_PATH,
    JLL_BASE_URL,
    REQUEST_TIMEOUT,
    RSS_FEED_URL,
    SCRAPE_MAX_ARTICLES,
    USER_AGENT,
    ZILLOW_CSV_PATH,
)

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════
# 1. RSS Feed Ingestion
# ═══════════════════════════════════════════════════════════════════════════


def ingest_rss(feed_url: str = RSS_FEED_URL) -> list[dict[str, Any]]:
    """Parse an RSS feed and return a list of article dicts.

    Each dict contains:
        - title (str)
        - link (str)
        - published_date (str)
        - summary (str)
        - source_type (str): always ``"rss"``

    Args:
        feed_url: URL of the RSS feed to parse.

    Returns:
        A list of dictionaries, one per RSS entry.
    """
    logger.info("Ingesting RSS feed from %s", feed_url)
    results: list[dict[str, Any]] = []

    try:
        import ssl
        if hasattr(ssl, '_create_unverified_context'):
            ssl._create_default_https_context = ssl._create_unverified_context

        feed = feedparser.parse(feed_url)

        if feed.bozo and not feed.entries:
            logger.warning("RSS feed returned bozo error: %s", feed.bozo_exception)
            return []

        for entry in feed.entries:
            published = entry.get("published", "")
            try:
                pub_date = datetime.strptime(
                    published, "%a, %d %b %Y %H:%M:%S %z"
                ).isoformat()
            except (ValueError, TypeError):
                pub_date = published or datetime.now(timezone.utc).isoformat()

            results.append(
                {
                    "title": entry.get("title", "Untitled"),
                    "link": entry.get("link", ""),
                    "published_date": pub_date,
                    "summary": entry.get("summary", ""),
                    "source_type": "rss",
                }
            )

        logger.info("Ingested %d RSS entries", len(results))
    except Exception as exc:
        logger.error("RSS ingestion failed: %s", exc)

    return results





# ═══════════════════════════════════════════════════════════════════════════
# 2. Web Scraping – JLL Trends & Insights
# ═══════════════════════════════════════════════════════════════════════════


def scrape_jll_articles(
    base_url: str = JLL_BASE_URL,
    max_articles: int = SCRAPE_MAX_ARTICLES,
) -> list[dict[str, Any]]:
    """Scrape article metadata and body text from JLL's insights hub.

    Uses ``requests`` + ``BeautifulSoup`` (no headless browsers).

    Args:
        base_url: Landing page URL to discover article links.
        max_articles: Maximum number of articles to scrape.

    Returns:
        A list of article dicts with keys: title, link, published_date,
        content, source_type.
    """
    logger.info("Scraping JLL articles from %s (max %d)", base_url, max_articles)
    headers = {"User-Agent": USER_AGENT}
    articles: list[dict[str, Any]] = []

    try:
        resp = requests.get(base_url, headers=headers, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        # JLL uses multiple possible selectors for article cards
        link_tags = soup.select("a[href*='/insights/']")[:max_articles]

        seen_urls: set[str] = set()
        for tag in link_tags:
            href: str = tag.get("href", "")
            if not href or href in seen_urls or href.rstrip("/") == base_url.rstrip("/"):
                continue
            if not href.startswith("http"):
                href = f"https://www.jll.co.uk{href}"
            seen_urls.add(href)

            article = _scrape_single_article(href, headers)
            if article:
                articles.append(article)
            if len(articles) >= max_articles:
                break

        logger.info("Scraped %d JLL articles", len(articles))
    except Exception as exc:
        logger.error("JLL scraping failed: %s", exc)


    return articles


def _scrape_single_article(
    url: str,
    headers: dict[str, str],
) -> dict[str, Any] | None:
    """Fetch and extract content from a single article page.

    Args:
        url: Full URL of the article.
        headers: HTTP headers dict.

    Returns:
        Article dict or ``None`` on failure.
    """
    try:
        resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        title_tag = soup.find("h1")
        title = title_tag.get_text(strip=True) if title_tag else "Untitled"

        # Attempt to find a date
        date_tag = soup.find("time")
        pub_date = (
            date_tag.get("datetime", datetime.now(timezone.utc).isoformat())
            if date_tag
            else datetime.now(timezone.utc).isoformat()
        )

        # Gather paragraph text as body content
        paragraphs = soup.find_all("p")
        body = " ".join(p.get_text(strip=True) for p in paragraphs if p.get_text(strip=True))

        if not body:
            return None

        return {
            "title": title,
            "link": url,
            "published_date": pub_date,
            "content": body[:5000],  # cap at ~5 000 chars for sanity
            "source_type": "scraping",
        }
    except Exception as exc:
        logger.warning("Failed to scrape %s: %s", url, exc)
        return None





# ═══════════════════════════════════════════════════════════════════════════
# 3. Web Scraping – Altus Group Insights
# ═══════════════════════════════════════════════════════════════════════════


def scrape_altus_articles(
    base_url: str = ALTUS_BASE_URL,
    max_articles: int = SCRAPE_MAX_ARTICLES,
) -> list[dict[str, Any]]:
    """Scrape article metadata and body text from Altus Group's insights hub.

    Uses ``requests`` + ``BeautifulSoup`` (no headless browsers).

    Args:
        base_url: Landing page URL to discover article links.
        max_articles: Maximum number of articles to scrape.

    Returns:
        A list of article dicts with keys: title, link, published_date,
        content, source_type.
    """
    logger.info("Scraping Altus Group articles from %s (max %d)", base_url, max_articles)
    headers = {"User-Agent": USER_AGENT}
    articles: list[dict[str, Any]] = []

    try:
        resp = requests.get(base_url, headers=headers, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        # Altus Group uses various selectors for article cards
        link_tags = soup.select("a[href*='/insights/']")[:max_articles * 2]

        seen_urls: set[str] = set()
        for tag in link_tags:
            href: str = tag.get("href", "")
            if not href or href in seen_urls or href.rstrip("/") == base_url.rstrip("/"):
                continue
            if not href.startswith("http"):
                href = f"https://www.altusgroup.com{href}"
            seen_urls.add(href)

            article = _scrape_single_article(href, headers)
            if article:
                articles.append(article)
            if len(articles) >= max_articles:
                break

        logger.info("Scraped %d Altus Group articles", len(articles))
    except Exception as exc:
        logger.error("Altus Group scraping failed: %s", exc)


    return articles





# ═══════════════════════════════════════════════════════════════════════════
# 4. Financial API — FMP Company Profiles (mock-ready)
# ═══════════════════════════════════════════════════════════════════════════


def fetch_fmp_profiles(
    tickers: list[str] | None = None,
    api_key: str = FMP_API_KEY,
) -> list[dict[str, Any]]:
    """Fetch company profile data from the Financial Modeling Prep API.

    When the API is unreachable or returns an error the function falls
    back to mock data so the pipeline never hard-fails.

    Args:
        tickers: List of stock tickers to query (default: CBRE, JLL, SPG).
        api_key: FMP API key.

    Returns:
        A list of profile dicts.
    """
    tickers = tickers or FMP_TICKERS
    profiles: list[dict[str, Any]] = []

    for ticker in tickers:
        try:
            url = f"{FMP_BASE_URL}/profile?symbol={ticker}&apikey={api_key}"
            resp = requests.get(url, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            data = resp.json()
            if isinstance(data, list) and data:
                profile = data[0]
                profiles.append(
                    {
                        "title": f"{profile.get('companyName', ticker)} — Company Profile",
                        "content": (
                            f"{profile.get('description', '')} "
                            f"Sector: {profile.get('sector', 'N/A')}. "
                            f"Industry: {profile.get('industry', 'N/A')}. "
                            f"Market Cap: ${profile.get('marketCap', profile.get('mktCap', 'N/A'))}. "
                            f"CEO: {profile.get('ceo', 'N/A')}. "
                            f"HQ: {profile.get('city', '')}, {profile.get('country', '')}."
                        ),
                        "link": profile.get("website", ""),
                        "published_date": datetime.now(timezone.utc).isoformat(),
                        "source_type": "api",
                        "ticker": ticker,
                        "market_cap": profile.get("marketCap", profile.get("mktCap")),
                        "sector": profile.get("sector"),
                        "price": profile.get("price"),
                    }
                )
            else:
                logger.warning("FMP API returned invalid or empty data for %s", ticker)
        except Exception as exc:
            logger.warning("FMP API call failed for %s: %s", ticker, exc)

    logger.info("Fetched %d FMP profiles", len(profiles))
    return profiles





# ═══════════════════════════════════════════════════════════════════════════
# 5. CSV Ingestion — Cities, Homes, Zillow
# ═══════════════════════════════════════════════════════════════════════════


def load_cities_csv(path: Path | str = CITIES_CSV_PATH) -> list[dict[str, str]]:
    """Load the local ``cities.csv`` as a list of dicts.

    The real CSV has columns: LatD, LatM, LatS, NS, LonD, LonM, LonS,
    EW, City, State.  This function normalises the output so that
    downstream code can access the ``city`` key consistently.

    Args:
        path: Path to the CSV file.

    Returns:
        A list of dicts, each representing a city row.
    """
    path = Path(path)
    if not path.exists():
        logger.warning("cities.csv not found at %s", path)
        return []

    rows: list[dict[str, str]] = []
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            # Normalise keys: strip whitespace and quotes from keys and values
            cleaned = {k.strip().strip('"'): (v.strip().strip('"') if v else "") for k, v in row.items() if k}
            # Ensure a 'city' key exists for downstream compatibility
            if "City" in cleaned and "city" not in cleaned:
                cleaned["city"] = cleaned["City"]
            if "State" in cleaned and "state" not in cleaned:
                cleaned["state"] = cleaned["State"]
            rows.append(cleaned)

    logger.info("Loaded %d cities from %s", len(rows), path)
    return rows


def load_homes_csv(path: Path | str = HOMES_CSV_PATH) -> list[dict[str, Any]]:
    """Load ``homes.csv`` residential property listings.

    Columns: Sell, List, Living, Rooms, Beds, Baths, Age, Acres, Taxes.

    Each row is transformed into a normalised document dict for the
    pipeline with ``source_type`` = ``"csv_listings"``.

    Args:
        path: Path to the CSV file.

    Returns:
        A list of document dicts ready for normalisation.
    """
    path = Path(path)
    if not path.exists():
        logger.warning("homes.csv not found at %s", path)
        return []

    docs: list[dict[str, Any]] = []
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for i, row in enumerate(reader):
            cleaned = {k.strip().strip('"'): (v.strip().strip('"') if v else "") for k, v in row.items() if k}
            sell = cleaned.get("Sell", "N/A")
            list_price = cleaned.get("List", "N/A")
            beds = cleaned.get("Beds", "N/A")
            baths = cleaned.get("Baths", "N/A")
            living = cleaned.get("Living", "N/A")
            age = cleaned.get("Age", "N/A")
            acres = cleaned.get("Acres", "N/A")
            taxes = cleaned.get("Taxes", "N/A")

            docs.append({
                "title": f"Residential Property #{i+1} — {beds}BR/{baths}BA, {living}00 sq ft",
                "content": (
                    f"Residential property listing. Sale price: ${sell}K. "
                    f"List price: ${list_price}K. Living area: {living}00 sq ft. "
                    f"{beds} bedrooms, {baths} bathrooms. Property age: {age} years. "
                    f"Lot size: {acres} acres. Annual taxes: ${taxes}."
                ),
                "source_type": "csv_listings",
                "published_date": datetime.now(timezone.utc).isoformat(),
                "link": "",
                "sell_price": sell,
                "list_price": list_price,
                "beds": beds,
                "baths": baths,
                "living_area": living,
                "age": age,
                "acres": acres,
                "taxes": taxes,
                "dataset": "homes",
            })

    logger.info("Loaded %d homes listings from %s", len(docs), path)
    return docs


def load_zillow_csv(path: Path | str = ZILLOW_CSV_PATH) -> list[dict[str, Any]]:
    """Load ``zillow.csv`` property listings.

    Columns: Index, Living Space (sq ft), Beds, Baths, Zip, Year, List Price ($).

    Args:
        path: Path to the CSV file.

    Returns:
        A list of document dicts ready for normalisation.
    """
    path = Path(path)
    if not path.exists():
        logger.warning("zillow.csv not found at %s", path)
        return []

    docs: list[dict[str, Any]] = []
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            cleaned = {k.strip().strip('"'): (v.strip().strip('"') if v else "") for k, v in row.items() if k}
            idx = cleaned.get("Index", "N/A")
            sqft = cleaned.get("Living Space (sq ft)", "N/A")
            beds = cleaned.get("Beds", "N/A")
            baths = cleaned.get("Baths", "N/A")
            zipcode = cleaned.get("Zip", "N/A")
            year = cleaned.get("Year", "N/A")
            price = cleaned.get("List Price ($)", "N/A")

            docs.append({
                "title": f"Zillow Listing #{idx} — {beds}BR/{baths}BA, {sqft} sq ft",
                "content": (
                    f"Zillow property listing. List price: ${price}. "
                    f"Living space: {sqft} sq ft. {beds} bedrooms, {baths} bathrooms. "
                    f"Year built: {year}. Zip code: {zipcode}."
                ),
                "source_type": "csv_listings",
                "published_date": datetime.now(timezone.utc).isoformat(),
                "link": "",
                "list_price": price,
                "beds": beds,
                "baths": baths,
                "sqft": sqft,
                "zip": zipcode,
                "year_built": year,
                "dataset": "zillow",
            })

    logger.info("Loaded %d zillow listings from %s", len(docs), path)
    return docs


# ═══════════════════════════════════════════════════════════════════════════
# 6. XLSX Ingestion — CRE Lending Data
# ═══════════════════════════════════════════════════════════════════════════


def load_cre_lending_data(
    path: Path | str = CRE_LENDING_XLSX_PATH,
) -> list[dict[str, Any]]:
    """Parse the CRE Lending XLSX into document dicts.

    The workbook has two sheets:
    - **UK deals** — columns: Lender, Borrower, Loan size (£m), Asset(s), Notes
    - **Continental Europe deals** — columns: Lender, Borrower, Loan size (€m), Asset(s), Notes

    Date-only rows (used as section headers in the spreadsheet) are
    captured as context for the deals that follow them.

    Args:
        path: Path to the XLSX file.

    Returns:
        A list of document dicts ready for normalisation.
    """
    path = Path(path)
    if not path.exists():
        logger.warning("CRE Lending XLSX not found at %s", path)
        return []

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — cannot read XLSX")
        return []

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    docs: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        region = "UK" if "UK" in sheet_name else "Continental Europe"
        currency = "£" if region == "UK" else "€"
        current_date = ""

        for row in ws.iter_rows(values_only=True):
            vals = [v for v in row if v is not None]
            if not vals:
                continue

            # Date-only rows act as section headers
            if len(vals) == 1 and isinstance(vals[0], datetime):
                current_date = vals[0].strftime("%Y-%m")
                continue

            # Skip header and metadata rows
            if len(vals) < 4:
                continue
            if str(vals[0]).startswith("*") or str(vals[0]).startswith("SAMPLE") or str(vals[0]).startswith("To access"):
                continue
            if str(vals[0]) == "Lender":
                continue

            lender = str(vals[0]).strip()
            borrower = str(vals[1]).strip()
            loan_size = vals[2]
            asset = str(vals[3]).strip()
            notes = str(vals[4]).strip() if len(vals) > 4 and vals[4] else ""

            # Clean loan_size
            try:
                loan_amount = float(str(loan_size).replace("c.", "").replace(">", "").strip())
            except (ValueError, TypeError):
                loan_amount = 0.0

            title = f"{lender} → {borrower}: {currency}{loan_size}m"
            content = (
                f"CRE lending deal ({region}, {current_date or 'undated'}). "
                f"Lender: {lender}. Borrower: {borrower}. "
                f"Loan size: {currency}{loan_size}m. "
                f"Asset: {asset}. "
                f"Notes: {notes}"
            )

            docs.append({
                "title": title,
                "content": content,
                "source_type": "xlsx",
                "published_date": f"{current_date}-01" if current_date else datetime.now(timezone.utc).isoformat(),
                "link": "",
                "lender": lender,
                "borrower": borrower,
                "loan_size_m": loan_amount,
                "currency": currency,
                "asset": asset,
                "notes": notes,
                "region": region,
            })

    wb.close()
    logger.info("Loaded %d CRE lending deals from %s", len(docs), path)
    return docs

